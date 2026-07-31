import sys
import os
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

ALLOWED_SOURCE_DC = [
    'JNP', 'MAU', 'ALG', 'SPR', 'MTH',
    'MZN', 'JHS', 'AYP', 'DEO', 'MRZ', 'RBR'
]

IDENTITY_COLS = 3
BLOCK_SIZE    = 6
IDX_OFD       = 0
IDX_FWD_TASK  = 1
IDX_FWD_1K    = 2
IDX_OFP       = 3
IDX_REV_TASK  = 4
IDX_REV_1K    = 5

C_FWD_TITLE   = "1E1B4B"
C_FWD_HDR     = "312E81"
C_REV_TITLE   = "581C87"
C_REV_HDR     = "6B21A8"
C_HDR_FONT    = "FFFFFF"
C_BORDER      = "CBD5E1"

CF_GREEN_BG   = "BBEFCF"
CF_GREEN_FONT = "166534"
CF_YELLOW_BG  = "FEF08A"
CF_YELLOW_FONT= "854D0E"
CF_RED_BG     = "FECACA"
CF_RED_FONT   = "991B1B"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(hex_color, bold=False, size=10):
    return Font(color=hex_color, bold=bold, size=size)

def _border(hex_color=C_BORDER):
    side = Side(style="thin", color=hex_color)
    return Border(left=side, right=side, top=side, bottom=side)

def _center():
    return Alignment(horizontal="center", vertical="center")

def _fmt_date(dt):
    if not isinstance(dt, datetime):
        return str(dt)
    months = ['Jan','Feb','Mar','Apr','May','Jun',
              'Jul','Aug','Sep','Oct','Nov','Dec']
    return f"{dt.day}-{months[dt.month-1]}-{dt.year}"

def _safe_float(val, fallback=0.0):
    try:
        return float(val) if val is not None else fallback
    except (TypeError, ValueError):
        return fallback

def parse_task_per_1k(ws):
    max_col = ws.max_column
    max_row = ws.max_row
    row1 = [ws.cell(1, c).value for c in range(1, max_col + 1)]

    block_starts = []
    for i in range(IDENTITY_COLS, len(row1)):
        v = row1[i]
        if v is not None:
            block_starts.append((i + 1, v))

    if not block_starts:
        raise ValueError("No date/WTD blocks found in Task_per_1k row 1.")

    raw_rows = []
    for r in range(3, max_row + 1):
        dc = ws.cell(r, 1).value
        if dc is None:
            continue
        dc = str(dc).strip()
        if not dc:
            continue
        region = ws.cell(r, 2).value
        city   = ws.cell(r, 3).value
        raw_rows.append((r, dc, region, city))

    blocks = []
    for col_start, label in block_starts:
        is_wtd = isinstance(label, str) and label.strip().upper() == 'WTD'
        block_rows = []
        for (r, dc, region, city) in raw_rows:
            if dc not in ALLOWED_SOURCE_DC:
                continue
            ofd      = _safe_float(ws.cell(r, col_start + IDX_OFD).value)
            fwd_task = _safe_float(ws.cell(r, col_start + IDX_FWD_TASK).value)
            fwd_1k   = _safe_float(ws.cell(r, col_start + IDX_FWD_1K).value)
            ofp      = _safe_float(ws.cell(r, col_start + IDX_OFP).value)
            rev_task = _safe_float(ws.cell(r, col_start + IDX_REV_TASK).value)
            rev_1k   = _safe_float(ws.cell(r, col_start + IDX_REV_1K).value)

            if ofd == 0 and fwd_task == 0 and ofp == 0 and rev_task == 0:
                continue

            block_rows.append({
                'dc': dc, 'region': region, 'city': city,
                'ofd': ofd, 'fwd_task': fwd_task, 'fwd_1k': fwd_1k,
                'ofp': ofp, 'rev_task': rev_task, 'rev_1k': rev_1k,
            })

        blocks.append({'label': label, 'is_wtd': is_wtd, 'rows': block_rows})

    return blocks

def select_daily_block(blocks):
    today     = datetime.now().date()
    yesterday = today - timedelta(days=1)
    daily_blocks = [b for b in blocks if not b['is_wtd']]

    for b in daily_blocks:
        lbl = b['label']
        if isinstance(lbl, datetime) and lbl.date() == yesterday:
            return b

    dated = [(b['label'].date(), b) for b in daily_blocks if isinstance(b['label'], datetime)]
    if dated:
        dated.sort(key=lambda x: x[0], reverse=True)
        return dated[0][1]

    if daily_blocks:
        return daily_blocks[0]

    raise ValueError("No daily date blocks found.")

def select_wtd_block(blocks):
    for b in blocks:
        if b['is_wtd']:
            return b
    if blocks:
        return blocks[-1]
    raise ValueError("No WTD block found.")

def build_date_range(blocks):
    dates = [b['label'] for b in blocks if not b['is_wtd'] and isinstance(b['label'], datetime)]
    if not dates:
        return ''
    dates.sort()
    if len(dates) == 1:
        return _fmt_date(dates[0])
    return f"{_fmt_date(dates[0])} - {_fmt_date(dates[-1])}"

def write_summary_sheet(wb, daily_block, wtd_block, date_range_str):
    ws = wb.create_sheet("SUMMARY")
    ws.sheet_view.showGridLines = False

    daily_date_str = _fmt_date(daily_block['label']) if isinstance(daily_block['label'], datetime) else str(daily_block['label'])

    FWD_HEADERS = ['Date', 'Source_DC', 'OFD', 'Forward_Task', 'Fwd_Task_per_1k']
    REV_HEADERS = ['Date', 'Source_DC', 'OFP', 'Reverse_Task', 'Rev_Task_per_1k']

    fwd_daily = sorted(daily_block['rows'], key=lambda r: r['fwd_1k'], reverse=True)
    rev_daily = sorted(daily_block['rows'], key=lambda r: r['rev_1k'], reverse=True)
    fwd_wtd   = sorted(wtd_block['rows'],   key=lambda r: r['fwd_1k'], reverse=True)
    rev_wtd   = sorted(wtd_block['rows'],   key=lambda r: r['rev_1k'], reverse=True)

    max_daily  = max(len(fwd_daily), len(rev_daily))
    max_weekly = max(len(fwd_wtd),   len(rev_wtd))

    output = []
    output.append(['Forward EI', '', '', '', '', '', 'Reverse EI', '', '', '', ''])
    output.append(FWD_HEADERS + [''] + REV_HEADERS)

    for i in range(max_daily):
        f = fwd_daily[i] if i < len(fwd_daily) else None
        r = rev_daily[i] if i < len(rev_daily) else None
        output.append([
            daily_date_str if f else '', f['dc'] if f else '',
            f['ofd'] if f else '', f['fwd_task'] if f else '',
            round(f['fwd_1k'], 2) if f else '',
            '',
            daily_date_str if r else '', r['dc'] if r else '',
            r['ofp'] if r else '', r['rev_task'] if r else '',
            round(r['rev_1k'], 2) if r else '',
        ])

    output.append([''] * 11)
    output.append([''] * 11)

    weekly_title_row_idx = len(output) + 1
    output.append(['Weekly Forward EI', '', '', '', '', '', 'Weekly Reverse EI', '', '', '', ''])
    weekly_hdr_row_idx = len(output) + 1
    output.append(FWD_HEADERS + [''] + REV_HEADERS)

    for i in range(max_weekly):
        f = fwd_wtd[i] if i < len(fwd_wtd) else None
        r = rev_wtd[i] if i < len(rev_wtd) else None
        output.append([
            date_range_str if f else '', f['dc'] if f else '',
            f['ofd'] if f else '', f['fwd_task'] if f else '',
            round(f['fwd_1k'], 2) if f else '',
            '',
            date_range_str if r else '', r['dc'] if r else '',
            r['ofp'] if r else '', r['rev_task'] if r else '',
            round(r['rev_1k'], 2) if r else '',
        ])

    for row_idx, row_data in enumerate(output, start=1):
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    total_rows = len(output)

    for r in range(1, total_rows + 1):
        for fmt_col, fmt in [(1,'@'),(2,'@'),(3,'0'),(4,'0'),(5,'0.00'),(7,'@'),(8,'@'),(9,'0'),(10,'0'),(11,'0.00')]:
            ws.cell(r, fmt_col).number_format = fmt
        for c in range(1, 12):
            ws.cell(r, c).alignment = _center()

    def _style_merged_title(row, c1, c2, bg):
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row, c1)
        cell.fill = _fill(bg)
        cell.font = _font(C_HDR_FONT, bold=True, size=11)
        cell.alignment = _center()

    def _style_sub_header(row, c1, c2, bg):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row, c)
            cell.fill = _fill(bg)
            cell.font = _font(C_HDR_FONT, bold=True)
            cell.alignment = _center()

    _style_merged_title(1, 1, 5, C_FWD_TITLE)
    _style_merged_title(1, 7, 11, C_REV_TITLE)
    _style_sub_header(2, 1, 5, C_FWD_HDR)
    _style_sub_header(2, 7, 11, C_REV_HDR)
    _style_merged_title(weekly_title_row_idx, 1, 5, C_FWD_TITLE)
    _style_merged_title(weekly_title_row_idx, 7, 11, C_REV_TITLE)
    _style_sub_header(weekly_hdr_row_idx, 1, 5, C_FWD_HDR)
    _style_sub_header(weekly_hdr_row_idx, 7, 11, C_REV_HDR)

    bd = _border(C_BORDER)
    def _apply_borders(rs, nr, cs, nc):
        for r in range(rs, rs + nr):
            for c in range(cs, cs + nc):
                ws.cell(r, c).border = bd

    if max_daily > 0:
        _apply_borders(1, 2 + max_daily, 1, 5)
        _apply_borders(1, 2 + max_daily, 7, 5)
    if max_weekly > 0:
        _apply_borders(weekly_title_row_idx, 2 + max_weekly, 1, 5)
        _apply_borders(weekly_title_row_idx, 2 + max_weekly, 7, 5)

    def _add_cf(col_letter, r_start, r_end, lo, hi):
        rng = f"{col_letter}{r_start}:{col_letter}{r_end}"
        ws.conditional_formatting.add(rng, CellIsRule(
            operator='lessThan', formula=[str(lo)],
            fill=_fill(CF_GREEN_BG), font=_font(CF_GREEN_FONT)))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator='between', formula=[str(lo), str(hi)],
            fill=_fill(CF_YELLOW_BG), font=_font(CF_YELLOW_FONT)))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator='greaterThan', formula=[str(hi)],
            fill=_fill(CF_RED_BG), font=_font(CF_RED_FONT)))

    if max_daily > 0:
        _add_cf('E', 3, 2 + max_daily, 2.5, 6.0)
        _add_cf('K', 3, 2 + max_daily, 6.1, 10.0)
    if max_weekly > 0:
        _add_cf('E', weekly_hdr_row_idx + 1, weekly_hdr_row_idx + max_weekly, 2.5, 6.0)
        _add_cf('K', weekly_hdr_row_idx + 1, weekly_hdr_row_idx + max_weekly, 6.1, 10.0)

    col_widths = {1:16, 2:8, 3:10, 4:14, 5:16, 6:3, 7:16, 8:8, 9:10, 10:14, 11:16}
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

def parse_raw_tab(ws):
    max_col = ws.max_column
    max_row = ws.max_row
    headers = [ws.cell(1, c).value for c in range(1, max_col + 1)]

    col_map = {}
    for i, h in enumerate(headers):
        if h is not None:
            col_map[str(h).strip().lower()] = i

    dc_idx = None
    for candidate in ['source_dc', 'source dc', 'dc']:
        if candidate in col_map:
            dc_idx = col_map[candidate]
            break

    track_idx = None
    for candidate in ['final_tracking_no', 'tracking_id', 'tracking id', 'waybill']:
        if candidate in col_map:
            track_idx = col_map[candidate]
            break

    fwd_agt_idx = col_map.get('fwd_agent name') or col_map.get('fwd agent') or col_map.get('fwd_agent')
    rev_agt_idx = col_map.get('rev_agent name') or col_map.get('rev agent') or col_map.get('rev_agent')

    if dc_idx is None:
        raise ValueError('Column Source_DC not found in Raw tab.')
    if track_idx is None:
        raise ValueError('Column Final_tracking_no / Tracking_ID not found in Raw tab.')

    all_data_rows   = []
    filt_data_rows  = []

    for row_tuple in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
        row = list(row_tuple)
        if not any(v for v in row):
            continue
        all_data_rows.append(row)
        dc = str(row[dc_idx] or '').strip()
        if dc in ALLOWED_SOURCE_DC:
            filt_data_rows.append(row)

    return headers, filt_data_rows, col_map, track_idx, fwd_agt_idx, rev_agt_idx, dc_idx

def write_filtered_dc_tab(wb, headers, filt_rows):
    ws = wb.create_sheet("Filtered_Source_DC")
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font       = Font(bold=True)
        cell.fill       = PatternFill("solid", fgColor="F1F5F9")
        cell.alignment  = _center()
    for r_idx, row in enumerate(filt_rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

def write_fwd_ei_tab(wb, headers, filt_rows, track_idx):
    ws = wb.create_sheet("FWD EI")
    fwd_rows = []
    for row in filt_rows:
        tno = str(row[track_idx] or '').strip().upper()
        if tno.startswith('MYSC') or tno.startswith('MYSP'):
            fwd_rows.append(row)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = Font(bold=True)
        cell.fill      = PatternFill("solid", fgColor="F1F5F9")
        cell.alignment = _center()
    for r_idx, row in enumerate(fwd_rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

def write_rev_ei_tab(wb, headers, filt_rows, track_idx):
    ws = wb.create_sheet("REVERSE EI")
    rev_rows = []
    for row in filt_rows:
        tno = str(row[track_idx] or '').strip().upper()
        if tno.startswith('MYSR'):
            rev_rows.append(row)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = Font(bold=True)
        cell.fill      = PatternFill("solid", fgColor="F1F5F9")
        cell.alignment = _center()
    for r_idx, row in enumerate(rev_rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

C_WARN_TITLE = "991B1B"
C_WARN_HDR   = "B91C1C"

def write_agent_summary_tab(wb, filt_rows, track_idx, fwd_agt_idx, rev_agt_idx, dc_idx):
    ws = wb.create_sheet("Agent Summary")
    ws.sheet_view.showGridLines = False

    counts = {}
    for row in filt_rows:
        dc  = str(row[dc_idx] or '').strip()
        if not dc:
            continue
        tno = str(row[track_idx] or '').strip().upper()
        if tno.startswith('MYSC') or tno.startswith('MYSP'):
            agent = str(row[fwd_agt_idx] or '').strip() if fwd_agt_idx is not None else ''
        elif tno.startswith('MYSR'):
            agent = str(row[rev_agt_idx] or '').strip() if rev_agt_idx is not None else ''
        else:
            agent = ''
        if not agent:
            agent = '#N/A'
        counts.setdefault(dc, {})
        counts[dc][agent] = counts[dc].get(agent, 0) + 1

    dc_order = ALLOWED_SOURCE_DC
    agent_rows     = []
    counselled_rows = []
    warned_rows    = []

    for dc in dc_order:
        if dc not in counts:
            continue
        for agent in sorted(counts[dc]):
            cnt = counts[dc][agent]
            la  = agent.lower()
            agent_rows.append([dc, agent, cnt])
            if cnt > 2 and la not in ('#n/a', 'n/a'):
                counselled_rows.append([dc, agent, cnt])
            if cnt > 5 and la not in ('#n/a', 'n/a'):
                warned_rows.append([dc, agent, cnt])

    counselled_rows.sort(key=lambda x: x[2], reverse=True)
    warned_rows.sort(key=lambda x: x[2], reverse=True)

    max_rows = max(len(agent_rows), len(counselled_rows), len(warned_rows), 1)

    output = []
    output.append(['Agent Summary', '', '', '', 'Agent to be counselled', '', '', '', 'Agents to be Warned', '', ''])
    output.append(['Source_DC', 'Agent Name', 'Count', '', 'Source_DC', 'Agent Name', 'Count', '', 'Source_DC', 'Agent Name', 'Count'])
    for i in range(max_rows):
        r1 = agent_rows[i]     if i < len(agent_rows)     else ['', '', '']
        r2 = counselled_rows[i] if i < len(counselled_rows) else ['', '', '']
        r3 = warned_rows[i]    if i < len(warned_rows)    else ['', '', '']
        output.append(r1 + [''] + r2 + [''] + r3)

    for r_idx, row in enumerate(output, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val).alignment = _center()

    if len(output) > 1:
        def _title(row, c1, c2, bg):
            ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
            cell = ws.cell(row, c1)
            cell.fill = _fill(bg)
            cell.font = _font(C_HDR_FONT, bold=True, size=11)
            cell.alignment = _center()

        def _subhdr(row, c1, c2, bg):
            for c in range(c1, c2 + 1):
                cell = ws.cell(row, c)
                cell.fill = _fill(bg)
                cell.font = _font(C_HDR_FONT, bold=True)
                cell.alignment = _center()

        _title(1, 1, 3, C_FWD_TITLE)
        _title(1, 5, 7, C_REV_TITLE)
        _title(1, 9, 11, C_WARN_TITLE)
        _subhdr(2, 1, 3, C_FWD_HDR)
        _subhdr(2, 5, 7, C_REV_HDR)
        _subhdr(2, 9, 11, C_WARN_HDR)

        bd = _border(C_BORDER)
        def _bdr(rs, nr, cs, nc):
            for r in range(rs, rs + nr):
                for c in range(cs, cs + nc):
                    ws.cell(r, c).border = bd

        data_rows = len(output)
        _bdr(1, data_rows, 1, 3)
        _bdr(1, data_rows, 5, 3)
        _bdr(1, data_rows, 9, 3)

    for c, w in [(1,10),(2,28),(3,8),(4,3),(5,10),(6,28),(7,8),(8,3),(9,10),(10,28),(11,8)]:
        ws.column_dimensions[get_column_letter(c)].width = w

def generate_ei_report(source_file_path: str, output_file_path: str) -> str:
    wb_src = openpyxl.load_workbook(source_file_path, data_only=True)

    if 'Task_per_1k' not in wb_src.sheetnames:
        raise ValueError("Sheet 'Task_per_1k' not found in source workbook")

    blocks = parse_task_per_1k(wb_src['Task_per_1k'])
    daily_block = select_daily_block(blocks)
    wtd_block   = select_wtd_block(blocks)
    date_range  = build_date_range(blocks)

    if 'Raw' not in wb_src.sheetnames:
        headers, filt_rows, col_map, track_idx, fwd_agt_idx, rev_agt_idx, dc_idx = [], [], {}, None, None, None, None
    else:
        headers, filt_rows, col_map, track_idx, fwd_agt_idx, rev_agt_idx, dc_idx = parse_raw_tab(wb_src['Raw'])

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    write_summary_sheet(wb_out, daily_block, wtd_block, date_range)

    if filt_rows:
        write_filtered_dc_tab(wb_out, headers, filt_rows)
        write_fwd_ei_tab(wb_out, headers, filt_rows, track_idx)
        write_rev_ei_tab(wb_out, headers, filt_rows, track_idx)
        write_agent_summary_tab(wb_out, filt_rows, track_idx, fwd_agt_idx, rev_agt_idx, dc_idx)

    wb_out.save(output_file_path)
    return output_file_path
